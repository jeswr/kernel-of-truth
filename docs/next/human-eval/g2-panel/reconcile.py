#!/usr/bin/env python3
"""
reconcile.py — ingest the two completed annotator workbooks, compute
inter-annotator agreement, drive blind adjudication, mint the reconciled
HUMAN GOLD, and re-score arms A0-A3 against it.

Two-step protocol (disagreements adjudicated blind BEFORE scoring, per the
frozen g2 record):

  STEP 1  kappa   python3 reconcile.py kappa workbooks/annotator-H1-DONE.xlsx \
                                             workbooks/annotator-H2-DONE.xlsx
          -> results/agreement.json  (Cohen's kappa 3-cat + binary, overall
             and per arm; practice check; probe false-satisfaction per
             annotator)
          -> results/adjudication-sheet.csv  (one row per disagreeing REAL
             row, STILL BLIND: row_id + statement + both raw answers +
             empty `consensus` column). The two annotators jointly re-judge
             these rows blind and fill `consensus` (yes/no/cannot-say).
             If they cannot agree, write cannot-say (scores 0 — conservative).

  STEP 2  score   python3 reconcile.py score workbooks/annotator-H1-DONE.xlsx \
                            workbooks/annotator-H2-DONE.xlsx \
                            results/adjudication-sheet-completed.csv
          -> results/human-gold-labels.jsonl  (one line per (arm,item):
             {"arm","id","gold","vacuous","sound"})
          -> results/human-gold-report.json + results/human-gold-report.md
             (per-arm sound counts, precision at fixed denominator 84,
             vacuity-zeroed, two-sided 95% Wilson, exact McNemar vs the
             HUMAN-rescored A0, kappa per arm; optional per-item comparison
             vs the archived LLM-proxy labels — safe only AFTER gold exists).

Accepts .xlsx (openpyxl) or .csv workbooks. Fails closed (ERR_*) on missing
rows, invalid answers, failed practice items, or an incomplete adjudication.

Scoring rules (mirroring the frozen ONT-TYPE-G2/1 plan sections 7.3-7.6):
  sound(arm,item) = 1 iff gold == "yes" AND the slot is non-vacuous in that
  arm; cannot-say / missing / vacuous = 0. Fixed denominator 84.
  A0 has no vacuity concept (all slots count). Probes and practice rows are
  instrument checks only and never enter gold.

STATUS: outputs are HUMAN-GOLD (the sole adoption authority per the frozen
g2-import envelope), but they do NOT by themselves flip any frozen verdict:
the g2 n_gold>=500 instrument gate stays unattainable on this corpus (ops
amendment pending) and any g2-import verdict change needs a new frozen run.
"""
import csv
import json
import math
import os
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
RESULTS = os.path.join(HERE, "results")
ROOT = os.path.abspath(os.path.join(HERE, "..", "..", "..", ".."))
VALID = ("yes", "no", "cannot-say")
ARMS = ("A0", "A1", "A2", "A3")
PROBE_FALSE_SAT_BAR = 0.30   # reused from the frozen LLM-proxy design; advisory here

# archived LLM-proxy labels — read ONLY in `score` mode, AFTER gold exists
PROXY_A0 = os.path.join(ROOT, "poc/g2/labels-proxy.jsonl")
PROXY_ONTG2 = os.path.join(ROOT, "poc/ontology-import-g2/labels-ontg2.jsonl")


def die(msg):
    sys.exit("ERR_" + msg)


def load_key():
    p = os.path.join(HERE, "blind-key.json")
    if not os.path.exists(p):
        die("NO_BLIND_KEY: run build_workbooks.py first")
    with open(p, encoding="utf-8") as f:
        return json.load(f)


def read_workbook(path):
    """Return {row_id: (answer, reason)} from a completed .xlsx or .csv."""
    rows = {}
    if path.lower().endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            die("OPENPYXL_MISSING: convert the workbook to CSV and retry")
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb["items"] if "items" in wb.sheetnames else wb.active
        it = ws.iter_rows(min_row=2, values_only=True)
    elif path.lower().endswith(".csv"):
        f = open(path, newline="", encoding="utf-8-sig")
        r = csv.reader(f)
        next(r)  # header
        it = r
    else:
        die(f"BAD_EXT: {path} (use .xlsx or .csv)")
    for row in it:
        if row is None or len(row) < 4:
            continue
        rid = (row[1] or "").strip() if row[1] else ""
        if not rid:
            continue
        ans = (str(row[3]).strip().lower() if row[3] is not None else "")
        reason = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ""
        rows[rid] = (ans, reason)
    return rows


def check_and_split(key, book, who):
    """Validate one completed workbook against the blind key; fail closed."""
    real, probe = {}, {}
    for rid, meta in key["rows"].items():
        if rid not in book:
            die(f"MISSING_ROW: {who} row {rid} absent from workbook")
        ans, _ = book[rid]
        if ans not in VALID:
            die(f"INVALID_ANSWER: {who} row {rid} = {ans!r} (need yes/no/cannot-say)")
        if meta["kind"] == "practice":
            if ans != meta["expected"]:
                die(f"PRACTICE_FAILED: {who} row {rid}: answered {ans!r}, "
                    f"expected {meta['expected']!r} — annotator must redo the "
                    "practice items before the run counts")
        elif meta["kind"] == "probe":
            probe[rid] = ans
        else:
            real[rid] = ans
    return real, probe


def kappa_from_pairs(pairs, cats):
    """Cohen's kappa over a list of (a,b) labels."""
    n = len(pairs)
    if n == 0:
        return None
    po = sum(1 for a, b in pairs if a == b) / n
    pe = 0.0
    for c in cats:
        pa = sum(1 for a, _ in pairs if a == c) / n
        pb = sum(1 for _, b in pairs if b == c) / n
        pe += pa * pb
    if pe >= 1.0:
        return 1.0
    return (po - pe) / (1 - pe)


def wilson(yes, n, z=1.96):
    # z=1.96 matches the pinned analysis/ontg2.py convention (line 46), so
    # human-gold intervals are directly comparable to the frozen proxy ones.
    p = yes / n
    d = 1 + z * z / n
    centre = (p + z * z / (2 * n)) / d
    half = z * math.sqrt(p * (1 - p) / n + z * z / (4 * n * n)) / d
    return centre - half, centre + half


def mcnemar_exact(b, c):
    """Two-sided exact McNemar p (binomial sign test on the discordant pairs)."""
    n = b + c
    if n == 0:
        return 1.0
    k = min(b, c)
    tail = sum(math.comb(n, i) for i in range(k + 1)) / 2.0 ** n
    return min(1.0, 2.0 * tail)


def expand(key, real_labels):
    """row-level labels -> per (arm,item) labels."""
    out = {}
    for rid, ans in real_labels.items():
        for arm, iid in key["rows"][rid]["members"]:
            out[(arm, iid)] = ans
    return out


def arm_pairs(key, ra, rb, arm):
    pairs = []
    ea, eb = expand(key, ra), expand(key, rb)
    for (a, i), va in ea.items():
        if a == arm:
            pairs.append((va, eb[(a, i)]))
    return pairs


def mode_kappa(h1_path, h2_path):
    key = load_key()
    os.makedirs(RESULTS, exist_ok=True)
    b1, b2 = read_workbook(h1_path), read_workbook(h2_path)
    r1, p1 = check_and_split(key, b1, "H1")
    r2, p2 = check_and_split(key, b2, "H2")

    fs1 = sum(1 for v in p1.values() if v == "yes")
    fs2 = sum(1 for v in p2.values() if v == "yes")
    all_pairs = [(r1[rid], r2[rid]) for rid in sorted(r1)]
    report = {
        "schema": "g2hg/1-agreement",
        "n_real_rows": len(r1),
        "practice": "both annotators passed 2/2",
        "probe_false_satisfaction": {
            "H1": {"yes_on_probes": fs1, "n": len(p1), "rate": fs1 / len(p1)},
            "H2": {"yes_on_probes": fs2, "n": len(p2), "rate": fs2 / len(p2)},
            "bar_advisory": PROBE_FALSE_SAT_BAR,
            "ok": fs1 / len(p1) <= PROBE_FALSE_SAT_BAR and fs2 / len(p2) <= PROBE_FALSE_SAT_BAR,
        },
        "kappa_3cat_overall_rows": kappa_from_pairs(all_pairs, VALID),
        "kappa_binary_yes_overall_rows": kappa_from_pairs(
            [("y" if a == "yes" else "n", "y" if b == "yes" else "n")
             for a, b in all_pairs], ("y", "n")),
        "kappa_3cat_per_arm_items": {
            arm: kappa_from_pairs(arm_pairs(key, r1, r2, arm), VALID) for arm in ARMS},
        "kappa_binary_yes_per_arm_items": {
            arm: kappa_from_pairs(
                [("y" if a == "yes" else "n", "y" if b == "yes" else "n")
                 for a, b in arm_pairs(key, r1, r2, arm)], ("y", "n")) for arm in ARMS},
        "raw_agreement_rows": sum(1 for a, b in all_pairs if a == b) / len(all_pairs),
        "n_disagreements_rows": sum(1 for a, b in all_pairs if a != b),
        "note": ("kappa here is HUMAN-pair agreement (the real thing the proxy "
                 "kappa only stood in for). Compare per-arm binary kappa with the "
                 "proxy pair's 0.527/0.430/0.286 (A1/A2/A3) to price the "
                 "soft-hedged-modality instrument concern."),
    }
    with open(os.path.join(RESULTS, "agreement.json"), "w", encoding="utf-8") as f:
        json.dump(report, f, indent=1, sort_keys=True)
        f.write("\n")

    # blind adjudication sheet: need the statement text -> re-read a workbook file
    texts = {}
    for src in (h1_path, h2_path):
        pass  # texts come from the original workbooks below
    # statements are in the workbooks themselves (col C); re-read raw
    texts = read_statements(h1_path)
    adj = os.path.join(RESULTS, "adjudication-sheet.csv")
    with open(adj, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["row_id", "statement", "H1_answer", "H2_answer",
                    "consensus (yes / no / cannot-say)", "adjudication note"])
        for rid in sorted(r1):
            if r1[rid] != r2[rid]:
                w.writerow([rid, texts.get(rid, ""), r1[rid], r2[rid], "", ""])
    print(json.dumps(report, indent=1, sort_keys=True))
    print(f"\nwrote {os.path.join(RESULTS, 'agreement.json')}")
    print(f"wrote {adj}  ({report['n_disagreements_rows']} rows to adjudicate, blind)")


def read_statements(path):
    out = {}
    if path.lower().endswith(".xlsx"):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb["items"] if "items" in wb.sheetnames else wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[1]:
                out[str(row[1]).strip()] = str(row[2])
    else:
        with open(path, newline="", encoding="utf-8-sig") as f:
            r = csv.reader(f)
            next(r)
            for row in r:
                if len(row) > 2 and row[1].strip():
                    out[row[1].strip()] = row[2]
    return out


def mode_score(h1_path, h2_path, adj_path):
    key = load_key()
    os.makedirs(RESULTS, exist_ok=True)
    b1, b2 = read_workbook(h1_path), read_workbook(h2_path)
    r1, _ = check_and_split(key, b1, "H1")
    r2, _ = check_and_split(key, b2, "H2")

    # adjudicated consensus for every disagreement
    adj = {}
    with open(adj_path, newline="", encoding="utf-8-sig") as f:
        r = csv.reader(f)
        next(r)
        for row in r:
            if not row or not row[0].strip():
                continue
            rid, cons = row[0].strip(), row[4].strip().lower()
            if cons not in VALID:
                die(f"ADJUDICATION_INVALID: row {rid} consensus {cons!r}")
            adj[rid] = cons
    gold_rows = {}
    for rid in sorted(r1):
        if r1[rid] == r2[rid]:
            gold_rows[rid] = r1[rid]
        elif rid in adj:
            gold_rows[rid] = adj[rid]
        else:
            die(f"ADJUDICATION_MISSING: disagreement row {rid} not in {adj_path}")

    gold = expand(key, gold_rows)          # (arm,item) -> yes/no/cannot-say
    vac = key["vacuous_by_arm_item"]
    ids = sorted({i for (a, i) in gold if a == "A0"})
    if len(ids) != 84:
        die(f"ITEM_COUNT: {len(ids)} != 84")

    # per-arm scoring, vacuity-zeroed, fixed denominator 84
    labels_path = os.path.join(RESULTS, "human-gold-labels.jsonl")
    per_arm = {}
    with open(labels_path, "w", encoding="utf-8") as lf:
        for arm in ARMS:
            sound = {}
            for i in ids:
                g = gold[(arm, i)]
                v = vac[f"{arm}|{i}"]
                s = 1 if (g == "yes" and not v) else 0
                sound[i] = s
                lf.write(json.dumps({"arm": arm, "id": i, "gold": g,
                                     "vacuous": v, "sound": s}, sort_keys=True) + "\n")
            y = sum(sound.values())
            lb, ub = wilson(y, 84)
            per_arm[arm] = {"sound": y, "precision": y / 84,
                            "wilson95_lb": lb, "wilson95_ub": ub,
                            "nonvacuous": sum(1 for i in ids if not vac[f"{arm}|{i}"]),
                            "_sound_by_id": sound}

    # McNemar vs HUMAN-rescored A0 (paired on the same 84 ids)
    base = per_arm["A0"]["_sound_by_id"]
    for arm in ("A1", "A2", "A3"):
        cur = per_arm[arm]["_sound_by_id"]
        b = sum(1 for i in ids if base[i] == 1 and cur[i] == 0)
        c = sum(1 for i in ids if base[i] == 0 and cur[i] == 1)
        per_arm[arm]["mcnemar_vs_A0_human"] = {
            "b_A0only": b, "c_armonly": c, "p_two_sided_exact": mcnemar_exact(b, c)}

    # optional: price the LLM proxy against the human gold (post-gold, safe)
    proxy_cmp = compare_proxy(gold, ids)

    report = {
        "schema": "g2hg/1-human-gold-report",
        "status": ("HUMAN-GOLD (sole adoption authority per the frozen g2-import "
                   "envelope). Does not by itself flip any frozen verdict: the g2 "
                   "n_gold>=500 gate stays unattainable on this corpus (ops "
                   "amendment pending) and any g2-import verdict change needs a "
                   "new frozen run."),
        "n_items": 84,
        "scoring": "vacuity-zeroed, fixed denominator 84; cannot-say/vacuous = 0",
        "arms": {a: {k: v for k, v in per_arm[a].items() if k != "_sound_by_id"}
                 for a in ARMS},
        "kappa_binary_yes_per_arm_items": {
            arm: kappa_from_pairs(
                [("y" if a == "yes" else "n", "y" if b == "yes" else "n")
                 for a, b in arm_pairs(key, r1, r2, arm)], ("y", "n")) for arm in ARMS},
        "gold_construction": ("agree -> label; disagree -> blind joint adjudication; "
                              "no consensus -> cannot-say (scores 0, conservative)"),
        "proxy_comparison_post_gold": proxy_cmp,
        "reference_proxy_scores_for_context": {
            "A0": "33/84 (frozen g2 primary judge-pA)",
            "A1": "55/84", "A2": "53/84", "A3": "57/84 (all PROVISIONAL-ON-LLM-PROXY)"},
    }
    rp = os.path.join(RESULTS, "human-gold-report.json")
    with open(rp, "w", encoding="utf-8") as f:
        json.dump(report, f, indent=1, sort_keys=True)
        f.write("\n")

    md = [
        "# g2 / g2-import human-gold re-score (84 frozen slots, arms A0-A3)", "",
        "**Status:** HUMAN-GOLD — the sole adoption authority per the frozen",
        "g2-import envelope. No frozen verdict flips without a new frozen run;",
        "the g2 n>=500 instrument gate remains unattainable on this corpus.", "",
        "| Arm | Sound /84 | Precision | Wilson 95% | McNemar vs A0-human (b,c,p) | kappa (human pair, binary) |",
        "|---|---|---|---|---|---|",
    ]
    for arm in ARMS:
        a = report["arms"][arm]
        mc = a.get("mcnemar_vs_A0_human")
        mcs = (f"{mc['b_A0only']}, {mc['c_armonly']}, {mc['p_two_sided_exact']:.2e}"
               if mc else "—")
        kap = report["kappa_binary_yes_per_arm_items"][arm]
        md.append(f"| {arm} | {a['sound']}/84 | {a['precision']:.4f} | "
                  f"[{a['wilson95_lb']:.4f}, {a['wilson95_ub']:.4f}] | {mcs} | "
                  f"{kap:.4f} |")
    md += ["", "Proxy-vs-human per-item agreement (post-gold pricing of the LLM proxy):", "```json",
           json.dumps(proxy_cmp, indent=1, sort_keys=True), "```", ""]
    with open(os.path.join(RESULTS, "human-gold-report.md"), "w", encoding="utf-8") as f:
        f.write("\n".join(md))

    print(json.dumps({k: v for k, v in report.items() if k != "proxy_comparison_post_gold"},
                     indent=1, sort_keys=True))
    print(f"\nwrote {labels_path}\nwrote {rp}\nwrote {os.path.join(RESULTS, 'human-gold-report.md')}")


def compare_proxy(gold, ids):
    """AFTER human gold exists, price the archived LLM-proxy labels against it.
    Read-only; skipped gracefully if the archives are absent."""
    out = {}
    try:
        with open(PROXY_A0, encoding="utf-8") as f:
            pa0 = {}
            for line in f:
                if line.strip():
                    r = json.loads(line)
                    if "id" in r and ("answer_pA" in r or "answer" in r):
                        pa0[r["id"]] = r.get("answer_pA", r.get("answer"))
        if pa0:
            agree = sum(1 for i in ids if i in pa0 and pa0[i] == gold[("A0", i)])
            out["A0_pA_vs_human_agree"] = f"{agree}/{sum(1 for i in ids if i in pa0)}"
    except (OSError, json.JSONDecodeError):
        out["A0"] = "proxy archive unavailable — skipped"
    try:
        cnt = {}
        with open(PROXY_ONTG2, encoding="utf-8") as f:
            for line in f:
                if not line.strip():
                    continue
                r = json.loads(line)
                arm = r["arm"].upper()
                k = f"{arm}_pA_vs_human"
                a, t = cnt.get(k, (0, 0))
                cnt[k] = (a + (1 if r["answer_pA"] == gold[(arm, r["id"])] else 0), t + 1)
        for k, (a, t) in sorted(cnt.items()):
            out[k + "_agree"] = f"{a}/{t}"
    except (OSError, json.JSONDecodeError, KeyError):
        out["A1-A3"] = "proxy archive unavailable — skipped"
    return out


if __name__ == "__main__":
    if len(sys.argv) >= 4 and sys.argv[1] == "kappa":
        mode_kappa(sys.argv[2], sys.argv[3])
    elif len(sys.argv) >= 5 and sys.argv[1] == "score":
        mode_score(sys.argv[2], sys.argv[3], sys.argv[4])
    else:
        sys.exit(__doc__)
