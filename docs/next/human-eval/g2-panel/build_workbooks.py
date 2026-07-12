#!/usr/bin/env python3
"""
build_workbooks.py — HUMAN-GOLD reconciliation package for the g2 / g2-import
84-item panel (steering read #6).

Builds two BLIND annotator workbooks over the SAME 84 frozen kernel-v0 slots
rendered under all four arms:

  A0  frozen hard-4-sort baseline      poc/g2/materials/items.jsonl
  A1  BFO-only soft anchors            poc/ontology-import-g2/materials/arm-a1-bfo.jsonl
  A2  A1 + SUMO 'Normally...'          poc/ontology-import-g2/materials/arm-a2-bfo-sumo.jsonl
  A3  A2 + FrameNet 'Typically...'     poc/ontology-import-g2/materials/arm-a3-bfo-sumo-framenet.jsonl

Design decisions (all disclosed in README.md):
  * BLIND: arm labels, rule/form/subject provenance, and every LLM-proxy label
    are stripped. Annotators see only an opaque row id + the statement text.
    The row->(arm,item) mapping lives in blind-key.json (MAINTAINER-ONLY;
    never sent to annotators).
  * DEDUP-BY-TEXT: identical renderings across arms (14 of 336 slot-renderings;
    mostly A2==A3 where FrameNet adds no clause) collapse to ONE row whose
    answer is copied to every member (arm,item) pair at reconciliation time.
    This cuts load and enforces intra-annotator consistency by construction.
  * PROBES: 20 deranged-sort probes (5 per arm, seed-pinned sample of the
    pinned 20-per-arm probe files), expected answer "no" by construction,
    interleaved blind. Instrument check only; never part of gold.
  * PRACTICE: the 2 pinned calibration items head each workbook, marked
    PRACTICE. Annotator must get both right (checked by reconcile.py).
  * ORDER: per-annotator seed-pinned shuffle, same rule as run-g2lp.py::
    _run_block — sorted(row_ids, key=sha256(seed + "|" + row_id)).
  * QUARANTINE: this script never copies any answer_* field anywhere.
    blind-key.json carries only structure + per-(arm,item) vacuity flags
    (needed for the frozen vacuity-zeroed scoring), no labels.

Deterministic given the pinned inputs. The CSV files are the canonical pinned
artifacts (sha256s in package-manifest.json); the .xlsx files are convenience
copies of the same rows (xlsx bytes are not guaranteed stable across openpyxl
versions).

Usage:  python3 build_workbooks.py          # writes into ./workbooks/
"""
import csv
import hashlib
import json
import os
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.abspath(os.path.join(HERE, "..", "..", "..", ".."))
OUT = os.path.join(HERE, "workbooks")

SCHEMA = "g2hg/1"          # human-gold package schema tag; part of every seed
BUILD_DATE = "20260712"    # pinned; part of every seed (do not auto-date)
ANNOTATORS = ["H1", "H2"]
N_PROBES_PER_ARM = 5

ARM_FILES = {
    "A0": os.path.join(ROOT, "poc/g2/materials/items.jsonl"),
    "A1": os.path.join(ROOT, "poc/ontology-import-g2/materials/arm-a1-bfo.jsonl"),
    "A2": os.path.join(ROOT, "poc/ontology-import-g2/materials/arm-a2-bfo-sumo.jsonl"),
    "A3": os.path.join(ROOT, "poc/ontology-import-g2/materials/arm-a3-bfo-sumo-framenet.jsonl"),
}
PROBE_FILES = {
    "A0": os.path.join(ROOT, "poc/g2/materials/probes.jsonl"),
    "A1": os.path.join(ROOT, "poc/ontology-import-g2/materials/probes-a1.jsonl"),
    "A2": os.path.join(ROOT, "poc/ontology-import-g2/materials/probes-a2.jsonl"),
    "A3": os.path.join(ROOT, "poc/ontology-import-g2/materials/probes-a3.jsonl"),
}
CALIBRATION_FILE = os.path.join(ROOT, "poc/g2/materials/calibration-items.jsonl")
LABELS_ONTG2 = os.path.join(ROOT, "poc/ontology-import-g2/labels-ontg2.jsonl")

FORBIDDEN_FIELDS = {"answer_pA", "answer_pB", "answer_pA_frozen", "answer_pB_frozen",
                    "scored_pA", "scored_pB", "answer", "expected_llm"}


def sha256_file(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def sha256_str(s):
    return hashlib.sha256(s.encode("utf-8")).hexdigest()


def load_jsonl(path):
    rows = []
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return rows


def seeded_order(ids, seed):
    """The run-g2lp.py::_run_block rule: sort by sha256(seed|id)."""
    return sorted(ids, key=lambda i: sha256_str(seed + "|" + i))


def main():
    os.makedirs(OUT, exist_ok=True)

    # ---- 1. load real items per arm; quarantine check -----------------------
    arm_items = {}   # arm -> {id: text}
    for arm, path in ARM_FILES.items():
        d = {}
        for r in load_jsonl(path):
            d[r["id"]] = r["item"]
        if len(d) != 84:
            sys.exit(f"ERR_ITEM_COUNT: {arm} has {len(d)} items, expected 84")
        arm_items[arm] = d
    ids84 = sorted(arm_items["A0"])
    for arm in ("A1", "A2", "A3"):
        if sorted(arm_items[arm]) != ids84:
            sys.exit(f"ERR_ID_MISMATCH: {arm} ids differ from A0")

    # ---- 2. vacuity flags per (arm,item) from the pinned labels file --------
    # (structure only — the answer/scored fields are NOT read)
    vacuous = {}  # "A1|g2:pi:000" -> bool ; A0 has no vacuity concept (all False)
    for r in load_jsonl(LABELS_ONTG2):
        vacuous[f"{r['arm'].upper()}|{r['id']}"] = bool(r["vacuous"])
    for i in ids84:
        vacuous[f"A0|{i}"] = False
    for arm in ("A1", "A2", "A3"):
        missing = [i for i in ids84 if f"{arm}|{i}" not in vacuous]
        if missing:
            sys.exit(f"ERR_VACUITY_MISSING: {arm} missing {len(missing)} flags")

    # ---- 3. dedup real rows by exact statement text --------------------------
    text_to_row = {}   # text -> row dict
    for arm in ("A0", "A1", "A2", "A3"):
        for i in ids84:
            text = arm_items[arm][i]
            if text not in text_to_row:
                rid = "Q" + sha256_str(SCHEMA + "|real|" + text)[:8]
                text_to_row[text] = {"row_id": rid, "kind": "real",
                                     "text": text, "members": []}
            text_to_row[text]["members"].append([arm, i])
    real_rows = list(text_to_row.values())
    rid_set = {r["row_id"] for r in real_rows}
    if len(rid_set) != len(real_rows):
        sys.exit("ERR_ROWID_COLLISION: real rows")

    # ---- 4. probe rows: seed-pinned sample of 5 per arm ---------------------
    probe_rows = []
    for arm, path in PROBE_FILES.items():
        probes = {r["id"]: r["item"] for r in load_jsonl(path)}
        if len(probes) != 20:
            sys.exit(f"ERR_PROBE_COUNT: {arm} has {len(probes)}")
        picked = seeded_order(sorted(probes), f"{SCHEMA}|probe-sample|{arm}|{BUILD_DATE}")[:N_PROBES_PER_ARM]
        for pid in picked:
            rid = "Q" + sha256_str(SCHEMA + "|probe|" + probes[pid])[:8]
            probe_rows.append({"row_id": rid, "kind": "probe", "expected": "no",
                               "text": probes[pid], "members": [[arm, pid]]})
    if len({r["row_id"] for r in probe_rows} | rid_set) != len(probe_rows) + len(real_rows):
        sys.exit("ERR_ROWID_COLLISION: probes")

    # ---- 5. practice rows ----------------------------------------------------
    practice_rows = []
    for r in load_jsonl(CALIBRATION_FILE):
        rid = "P" + sha256_str(SCHEMA + "|practice|" + r["item"])[:8]
        practice_rows.append({"row_id": rid, "kind": "practice",
                              "expected": r["expected"], "text": r["item"],
                              "members": [["CAL", r["id"]]]})

    scored_rows = real_rows + probe_rows
    by_rid = {r["row_id"]: r for r in scored_rows + practice_rows}

    # ---- 6. per-annotator seed-pinned orders ---------------------------------
    orders = {}
    for ann in ANNOTATORS:
        seed = f"{SCHEMA}|{ann}|{BUILD_DATE}"
        orders[ann] = seeded_order([r["row_id"] for r in scored_rows], seed)

    # ---- 7. write CSV (canonical) + XLSX (convenience) -----------------------
    header = ["n", "row_id", "statement",
              "answer (yes / no / cannot-say)", "reason (short, your own words)"]
    csv_paths, xlsx_paths = {}, {}
    for ann in ANNOTATORS:
        rows_out = []
        n = 0
        for pr in practice_rows:
            n += 1
            rows_out.append([f"PRACTICE-{n}", pr["row_id"], pr["text"], "", ""])
        n = 0
        for rid in orders[ann]:
            n += 1
            rows_out.append([str(n), rid, by_rid[rid]["text"], "", ""])

        cpath = os.path.join(OUT, f"annotator-{ann}.csv")
        with open(cpath, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(header)
            w.writerows(rows_out)
        csv_paths[ann] = cpath

        xpath = os.path.join(OUT, f"annotator-{ann}.xlsx")
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.worksheet.datavalidation import DataValidation
            wb = Workbook()
            ws = wb.active
            ws.title = "items"
            ws.append(header)
            for c in ws[1]:
                c.font = Font(bold=True)
                c.fill = PatternFill("solid", fgColor="DDDDDD")
            for row in rows_out:
                ws.append(row)
            dv = DataValidation(type="list", formula1='"yes,no,cannot-say"',
                                allow_blank=True, showErrorMessage=True,
                                errorTitle="Invalid answer",
                                error="Use exactly: yes, no, or cannot-say")
            ws.add_data_validation(dv)
            dv.add(f"D2:D{ws.max_row}")
            ws.column_dimensions["A"].width = 12
            ws.column_dimensions["B"].width = 12
            ws.column_dimensions["C"].width = 100
            ws.column_dimensions["D"].width = 26
            ws.column_dimensions["E"].width = 50
            wrap = Alignment(wrap_text=True, vertical="top")
            for r in ws.iter_rows(min_row=2):
                r[2].alignment = wrap
                r[4].alignment = wrap
            ws.freeze_panes = "A2"
            ws2 = wb.create_sheet("READ ME FIRST")
            ws2.column_dimensions["A"].width = 110
            ws2["A1"] = ("Read INSTRUCTIONS-annotator.md before starting. "
                         "Judge each statement ONLY by the ordinary meaning of the quoted concept. "
                         "Answer every row: yes / no / cannot-say. The two PRACTICE rows come first. "
                         "Work alone; no lookups, no tools, no discussion with anyone.")
            ws2["A1"].alignment = wrap
            wb.save(xpath)
            xlsx_paths[ann] = xpath
        except ImportError:
            print("openpyxl not available — CSV-only build (see README fallback)")

    # ---- 8. blind key (MAINTAINER-ONLY) + manifest ---------------------------
    key = {
        "schema": SCHEMA + "-blindkey",
        "build_date": BUILD_DATE,
        "SEALED": "MAINTAINER-ONLY. Never send this file (or any part of it) to an annotator.",
        "n_real_rows": len(real_rows),
        "n_probe_rows": len(probe_rows),
        "n_practice_rows": len(practice_rows),
        "order_rule": "sorted(row_ids, key=sha256(seed|row_id)); seeds below",
        "order_seeds": {ann: f"{SCHEMA}|{ann}|{BUILD_DATE}" for ann in ANNOTATORS},
        "orders": orders,
        "rows": {r["row_id"]: {k: r[k] for k in r if k != "text"}
                 for r in scored_rows + practice_rows},
        "vacuous_by_arm_item": vacuous,
    }
    key_path = os.path.join(HERE, "blind-key.json")
    with open(key_path, "w", encoding="utf-8") as f:
        json.dump(key, f, indent=1, sort_keys=True)
        f.write("\n")

    manifest = {
        "schema": SCHEMA + "-manifest",
        "build_date": BUILD_DATE,
        "inputs_sha256": {os.path.relpath(p, ROOT): sha256_file(p)
                          for p in list(ARM_FILES.values()) + list(PROBE_FILES.values())
                          + [CALIBRATION_FILE, LABELS_ONTG2]},
        "outputs_sha256": {os.path.relpath(p, HERE): sha256_file(p)
                           for p in sorted(list(csv_paths.values())
                                           + list(xlsx_paths.values()) + [key_path])},
        "counts": {"real_rows_deduped": len(real_rows),
                   "slot_renderings_covered": sum(len(r["members"]) for r in real_rows),
                   "probe_rows": len(probe_rows), "practice_rows": len(practice_rows),
                   "rows_per_workbook": len(scored_rows) + len(practice_rows)},
        "note_xlsx": "CSV files are the canonical pinned artifacts; xlsx are convenience copies of the same rows.",
    }
    man_path = os.path.join(HERE, "package-manifest.json")
    with open(man_path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=1, sort_keys=True)
        f.write("\n")

    print(f"real rows (deduped): {len(real_rows)}  covering "
          f"{sum(len(r['members']) for r in real_rows)}/336 slot-renderings")
    print(f"probe rows: {len(probe_rows)}   practice rows: {len(practice_rows)}")
    print(f"rows per workbook (excl. header): {len(scored_rows) + len(practice_rows)}")
    for ann in ANNOTATORS:
        print(f"  {ann}: {csv_paths[ann]}" + (f" + {xlsx_paths[ann]}" if ann in xlsx_paths else ""))
    print(f"blind key (SEALED): {key_path}")
    print(f"manifest: {man_path}")


if __name__ == "__main__":
    main()
