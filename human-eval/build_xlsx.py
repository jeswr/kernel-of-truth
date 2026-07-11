#!/usr/bin/env python3
"""Build the self-contained human-eval workbook from human-eval-spec.json."""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

SPEC = "/tmp/claude-1000/-home-ec2-user-css/85798a0b-4e71-4020-b0b9-ac1fed9631d0/scratchpad/human-eval-spec.json"
OUT = "/tmp/claude-1000/-home-ec2-user-css/85798a0b-4e71-4020-b0b9-ac1fed9631d0/scratchpad/kernel-of-truth-human-eval-tasks.xlsx"
spec = json.load(open(SPEC))
T = {t["id"]: t for t in spec["tasks"]}

HDR = PatternFill("solid", fgColor="1F3864")
HDRF = Font(bold=True, color="FFFFFF")
INPUT = PatternFill("solid", fgColor="FFF2CC")   # yellow = you fill this in
CAL = PatternFill("solid", fgColor="E2EFDA")      # green = calibration (answer shown)
TITLE = Font(bold=True, size=14, color="1F3864")
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

def sheet(name):
    ws = wb.create_sheet(name[:31])
    return ws

def jl(x, sep="\n"):
    if isinstance(x, list): return sep.join(str(i) for i in x)
    if isinstance(x, dict): return sep.join(f"{k} = {v}" for k, v in x.items())
    return "" if x is None else str(x)

def instr_block(ws, task, extra=""):
    """Row 1 title, row 2 instructions block (merged, wrapped)."""
    ncol = 12
    ws.cell(1, 1, task["title"]).font = TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    txt = "HUMAN ROLE: " + jl(task.get("human_role")) + "\n\nINSTRUCTIONS:\n" + jl(task.get("instructions"))
    if task.get("n_annotators", 1) >= 2:
        txt += ("\n\n** THIS TASK NEEDS %d INDEPENDENT ANNOTATORS. ** Each annotator fills their OWN answer column "
                "(A1 / A2) WITHOUT looking at the other's answers — ideally give each person a separate copy of this file." % task["n_annotators"])
    if extra: txt += "\n\n" + extra
    c = ws.cell(2, 1, txt); c.alignment = WRAP
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
    ws.row_dimensions[2].height = 220
    return 4  # first table row

def dv(ws, col_letter, allowed, first, last):
    d = DataValidation(type="list", formula1='"%s"' % ",".join(allowed), allow_blank=True)
    ws.add_data_validation(d); d.add("%s%d:%s%d" % (col_letter, first, col_letter, last))

def write_table(ws, start, headers, rows, input_cols, widths, enum_cols=None, cal_count=0):
    """headers: list; rows: list of (cells, is_cal); input_cols: set of col idx (1-based) to shade yellow; enum_cols: {colidx:allowed}"""
    enum_cols = enum_cols or {}
    for j, h in enumerate(headers, 1):
        c = ws.cell(start, j, h); c.fill = HDR; c.font = HDRF; c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center"); c.border = BORDER
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = ws.cell(start + 1, 1)
    r = start + 1
    for cells, is_cal in rows:
        for j, v in enumerate(cells, 1):
            c = ws.cell(r, j, v); c.alignment = WRAP; c.border = BORDER
            if j in input_cols: c.fill = CAL if is_cal else INPUT
        r += 1
    last = r - 1
    for cidx, allowed in enum_cols.items():
        dv(ws, get_column_letter(cidx), allowed, start + 1, last)
    return last

# ---------------- g3 Pass A ----------------
t = T["g3.annotate"]
ws = sheet("g3 - Pass A")
top = instr_block(ws, t, "PASS A ONLY on this sheet. Do NOT open the Pass B sheet until you have finished ALL of Pass A "
                         "(seeing the condition lists would bias your ordinary-usage judgment). For each item read the situation, "
                         "then answer q1: under YOUR ordinary use of the key word, is the target claim true of the situation?")
rows = []
for it in t["items"]:
    rows.append(([it["instance_id"], it["text"], it["pass_a"]["target"], "", ""], False))
write_table(ws, top, ["instance_id", "Situation", "Target claim (q1: true of the situation?)", "Annotator A1: q1", "Annotator A2: q1"],
            rows, {4, 5}, [14, 70, 40, 16, 16], enum_cols={4: ["yes","no","cannot-say"], 5: ["yes","no","cannot-say"]})

# ---------------- g3 Pass B ----------------
ws = sheet("g3 - Pass B")
top = instr_block(ws, t, "PASS B. For each item: do the conditions (with referents fixed by the bindings) ALL hold in the situation? "
                         "Answer q2 = yes only if EVERY listed condition holds. If q2 = no, list the failing condition ids (e.g. c1,c3) in the next column.")
rows = []
for it in t["items"]:
    pb = it["pass_b"]
    conds = "\n".join("%s: %s" % (c["cid"], c["text"]) for c in pb.get("conditions", []))
    binds = jl(pb.get("bindings"))
    rows.append(([it["instance_id"], it["text"], binds, conds, "", "", "", ""], False))
write_table(ws, top,
            ["instance_id", "Situation", "Bindings (referents)", "Conditions", "A1: q2", "A1: failing cids", "A2: q2", "A2: failing cids"],
            rows, {5, 6, 7, 8}, [14, 55, 26, 46, 12, 16, 12, 16],
            enum_cols={5: ["yes","no","cannot-say"], 7: ["yes","no","cannot-say"]})

# ---------------- m0a ----------------
t = T["m0a"]
ws = sheet("m0a")
inv_note = "For 'inventory' items, choose the matching entry NUMBER from the 'm0a - shared inventory' sheet (or 'none')."
top = instr_block(ws, t, jl(t.get("_rendering_note")) + "\n\n" + inv_note +
                  "\n\nANSWER FORMAT depends on question_type: single-sense -> correct / incorrect / unclear; "
                  "candidate-list -> the NUMBER of the best-fitting candidate (see 'Candidate senses' cell) or 'none'; "
                  "inventory -> the NUMBER from the shared-inventory sheet or 'none'. Green rows are CALIBRATION (expected answer shown) — do them first to check yourself.")
rows = []
for cal in t.get("calibration", []):
    cs = "\n".join("%d: %s" % (i+1, s) for i, s in enumerate(cal.get("candidate_senses", []))) if cal.get("candidate_senses") else ""
    rows.append(([cal.get("id"), "CALIBRATION", cal.get("question_type"), cal.get("passage"), cal.get("marked_word"),
                  cal.get("proposed_sense",""), cs, "expected: %s" % cal.get("expected_answer",""), ""], True))
for it in t["items"]:
    cs = "\n".join("%d: %s" % (i+1, s) for i, s in enumerate(it.get("candidate_senses", []))) if it.get("candidate_senses") else ""
    rows.append(([it["itemId"], "", it.get("question_type"), it.get("passage"), it.get("marked_word"),
                  it.get("proposed_sense",""), cs, "", ""], False))
write_table(ws, top,
            ["itemId", "", "question_type", "Passage ([[marked word]])", "Marked word", "Proposed sense (single-sense)", "Candidate senses (candidate-list)", "answer", "note (optional)"],
            rows, {8, 9}, [12, 12, 12, 60, 14, 40, 40, 14, 24])

# m0a shared inventory  (spec stores it as a newline-delimited, self-numbered string: "1. afraid -- ...")
inv = t.get("shared_sense_inventory") or []
if isinstance(inv, str):
    inv = [l.strip() for l in inv.split("\n") if l.strip()]
ws2 = sheet("m0a - shared inventory")
ws2.cell(1,1,"m0a shared-sense inventory — for the 'inventory' items on the m0a sheet, pick the matching entry NUMBER (or 'none'). Entries are self-numbered.").font = Font(bold=True)
ws2.merge_cells(start_row=1,start_column=1,end_row=1,end_column=1)
ws2.cell(3,1,"shared-sense inventory entry").fill=HDR; ws2.cell(3,1).font=HDRF
ws2.column_dimensions["A"].width=120
for i, e in enumerate(inv, 1):
    c=ws2.cell(3+i,1, e if isinstance(e,str) else json.dumps(e)); c.alignment=WRAP

# ---------------- g9 review ----------------
t = T["g9.review"]
ws = sheet("g9 - review")
top = instr_block(ws, t, "For each sheet answer TWO yes/no/cannot-say questions about the proposed paraphrase (candidate_explication): "
                         "(1) substitutable — can it stand in for the word in every example and fill every <UNK> gap, meaning preserved (judge meaning not style)? "
                         "(2) cross_translatable — does it use only simple, broadly cross-linguistic words? Green rows are CALIBRATION.")
rows = []
def g9row(it, is_cal):
    return ([it.get("sheet_id"), it.get("word"), it.get("def"), it.get("syn",""),
             jl(it.get("examples")), jl(it.get("ambig_examples")), it.get("candidate_explication"),
             ("expected: %s" % it.get("expected_substitutable","")) if is_cal else "",
             ("expected: %s" % it.get("expected_cross_translatable","")) if is_cal else ""], is_cal)
for cal in t.get("calibration", []): rows.append(g9row(cal, True))
for it in t["items"]: rows.append(g9row(it, False))
write_table(ws, top,
            ["sheet_id", "word", "dictionary def", "sense id", "examples", "ambiguous examples (<UNK> = word removed)", "candidate_explication (paraphrase)", "substitutable", "cross_translatable"],
            rows, {8, 9}, [9, 14, 34, 12, 40, 55, 44, 16, 16],
            enum_cols={8: ["yes","no","cannot-say"], 9: ["yes","no","cannot-say"]})

# ---------------- g2 (blocked) ----------------
t = T["g2"]
ws = sheet("g2 (BLOCKED)")
ws.cell(1,1,t["title"]).font=TITLE; ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=6)
c=ws.cell(2,1,"** BLOCKED — no items yet. **\n\n"+jl(t.get("instructions"))+"\n\nHUMAN ROLE (for when it unblocks): "+jl(t.get("human_role"))+
          "\n\nSTATUS: "+jl(t.get("_status_note",""))); c.alignment=WRAP
ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=6); ws.row_dimensions[2].height=260
for j,w in enumerate([10,10,10,10,10,10],1): ws.column_dimensions[get_column_letter(j)].width=w

# ---------------- README (first) ----------------
ws = wb.active; ws.title = "README"
ws.cell(1,1,"Kernel of Truth — remaining human-evaluation tasks").font = Font(bold=True, size=16, color="1F3864")
overview = (
 "This workbook contains every remaining human judge / annotator task the Kernel of Truth programme needs, "
 "one task per tab. It is self-contained — you do NOT need any repository access.\n\n"
 "HOW TO USE:\n"
 " • Yellow cells are for you to fill in. Green cells are CALIBRATION items with the expected answer shown — do these first to check your understanding.\n"
 " • Enum cells have a dropdown (yes / no / cannot-say etc.).\n"
 " • These are BLIND tasks: judge each item only on what is shown; do not look for patterns across items; there is no hidden 'key' in the sheet.\n"
 " • Tasks needing 2 independent annotators have two answer columns (A1/A2) — the two people must NOT see each other's answers (ideally use separate copies of this file).\n\n"
 "TASKS:\n"
 " • g3 – Pass A / Pass B  — 2 independent annotators, 200 situations. Complete ALL of Pass A before opening Pass B.\n"
 " • m0a                   — 1 annotator (a 2nd is welcome for agreement), 300 word-sense items + shared-inventory tab.\n"
 " • g9 – review           — 1 annotator, 50 paraphrase-review sheets.\n"
 " • g2 (BLOCKED)          — a real task whose items don't exist yet (pending an upstream tool); included for reference only.\n\n"
 "NOTE: the separate 360-question judge task is already being done by another human and is intentionally NOT in this workbook.\n\n"
 "Generated 2026-07-11 from the frozen experiment records. Item text is copied byte-exact and gold answers are excluded (blindness-verified)."
)
c = ws.cell(3,1,overview); c.alignment = WRAP
ws.merge_cells(start_row=3,start_column=1,end_row=3,end_column=8); ws.row_dimensions[3].height=430
ws.column_dimensions["A"].width = 20
# order: README first, then tasks
order = ["README","g3 - Pass A","g3 - Pass B","m0a","m0a - shared inventory","g9 - review","g2 (BLOCKED)"]
wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)

wb.save(OUT)
print("wrote", OUT, os.path.getsize(OUT), "bytes")
print("tabs:", [s.title for s in wb._sheets])
