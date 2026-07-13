#!/usr/bin/env python3
"""Build the CURRENT human-review workbook (2026-07-13).

What the maintainer actually needs human eyes on NOW, curated honestly:
  - PRIMARY (needed): knull v4 blind STYLE spot-check (issue #31) — a maintainer sign-off
    step by design; unblocks the knull GPU control-arm run. Not proxiable.
  - OPTIONAL (only if a human is available): a small reconciliation sample of the g3
    proxy-graded leg. The affirmative legs are proxy-run (GPT-5.6/Fable = working gold per
    the 2026-07-12 no-bottleneck directive); this only validates the proxy, blocks nothing.

Supersedes the Jul-11 kernel-of-truth-human-eval-tasks.xlsx for "what is still open".
"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

REPO = "/home/ec2-user/css/kernel/kernel-of-truth"
SHEET_LINK = "https://docs.google.com/spreadsheets/d/1q_fBylZSctlCr7UkvAbiMA1AZxKFKvQ3"

H = Font(bold=True, size=12, color="FFFFFF")
HFILL = PatternFill("solid", fgColor="2F5597")
SUB = Font(bold=True, size=11)
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")
THIN = Border(*[Side(style="thin", color="D0D0D0")] * 4)
RESP = PatternFill("solid", fgColor="FFF2CC")  # response cells to fill = pale yellow

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = H; cell.fill = HFILL; cell.alignment = WRAP

def widths(ws, ws_widths):
    for col, w in ws_widths.items():
        ws.column_dimensions[col].width = w

wb = Workbook()

# ---------------- Tab 1: READ ME FIRST ----------------
ws = wb.active
ws.title = "READ ME FIRST"
ws.sheet_view.showGridLines = False
rows = [
    ("Kernel of Truth — current human-review needs", 16, True, "1F3864"),
    ("Prepared 2026-07-13 by the coordinator (Kern). Answers: \"Do you still need the human checks?\"", 11, False, None),
    ("", 11, False, None),
    ("SHORT ANSWER: mostly no. The affirmative annotation legs (g2 Π-soundness, g3 necessity/sufficiency,", 11, False, None),
    ("g9 review, m0a coverage) were all run with GPT-5.6 / Fable standing in for the human annotator, per your", 11, False, None),
    ("2026-07-12 directive (\"we may not get more human grading beyond the current sheet; use GPT-5.6/Fable as", 11, False, None),
    ("stand-in and get on with the work\"). Their verdicts are registered; proxy grading is the WORKING gold.", 11, False, None),
    ("I am NOT bottlenecking on human gold.", 11, False, None),
    ("", 11, False, None),
    ("The ONE thing that genuinely needs HUMAN eyes now (by design, not proxiable):", 12, True, "C00000"),
    ("  • Tab \"1. knull STYLE (NEEDED)\" — a blind STYLE judgement of 10 plain-dictionary definitions", 11, False, None),
    ("    (the knull control arm). This is a maintainer sign-off step (design §2.3, issue #31); it unblocks", 11, False, None),
    ("    the knull GPU run. ~5–10 minutes.", 11, False, None),
    ("", 11, False, None),
    ("OPTIONAL (only if a human is ever available — blocks nothing):", 12, True, "7F6000"),
    ("  • Tab \"2. OPTIONAL g3 reconcile\" — a 15-item sample of the g3 proxy-graded leg. Lets me report", 11, False, None),
    ("    proxy-vs-human agreement and re-run only if they materially differ. The g3 verdict currently rests", 11, False, None),
    ("    on the GPT-5.6/Fable proxy; this would upgrade a sample of it to human gold. Not blocking.", 11, False, None),
    ("", 11, False, None),
    ("Where the human annotator previously got to (do NOT re-do items already done there):", 12, True, "1F3864"),
    ("  " + SHEET_LINK, 11, False, None),
    ("  This workbook is the fresh, current curated set — it deliberately does NOT re-request the g2/g3/g9/m0a", 11, False, None),
    ("  full annotation portfolios (proxy-covered). If you'd rather I add any of those back, say so.", 11, False, None),
    ("", 11, False, None),
    ("Priorities:", 12, True, "1F3864"),
]
r = 1
for text, size, bold, color in rows:
    c = ws.cell(row=r, column=1, value=text)
    c.font = Font(bold=bold, size=size, color=(color or "000000"))
    r += 1
# priority table
ws.cell(row=r, column=1, value="Item").font = SUB
ws.cell(row=r, column=2, value="Priority").font = SUB
ws.cell(row=r, column=3, value="Blocks").font = SUB
r += 1
for item, pri, blocks in [
    ("knull v4 blind STYLE spot-check (Tab 1)", "NEEDED", "knull GPU control-arm run (issue #31 / d0hq → 1np7)"),
    ("g3 proxy reconciliation sample (Tab 2)", "optional", "nothing — validates the proxy gold only"),
]:
    ws.cell(row=r, column=1, value=item).alignment = TOP
    ws.cell(row=r, column=2, value=pri).alignment = TOP
    ws.cell(row=r, column=3, value=blocks).alignment = WRAP
    r += 1
widths(ws, {"A": 62, "B": 14, "C": 60})

# ---------------- Tab 2: knull STYLE (NEEDED) ----------------
spot = json.load(open(f"{REPO}/poc/knull/inputs-v3/plain-spotcheck-current.json"))
ws = wb.create_sheet("1. knull STYLE (NEEDED)")
ws.sheet_view.showGridLines = False
ws.cell(row=1, column=1, value="knull v4 plain store — BLIND style spot-check (do NOT open the 'knull KEY' tab until you have recorded all 10)").font = Font(bold=True, size=12, color="C00000")
ws.cell(row=2, column=1, value=spot["instructions"]).alignment = WRAP
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
ws.row_dimensions[2].height = 78
hdr = ["#", "Definition (read as if from a general dictionary)",
       "(a) ordinary dictionary-register English, to the scholarly standard? [yes / borderline / no]",
       "(b) fair definition a general dictionary could print? [yes / borderline / no]",
       "Notes (what, if anything, reads off)"]
hr = 4
for c, t in enumerate(hdr, 1):
    ws.cell(row=hr, column=c, value=t)
style_header(ws, hr, len(hdr))
for i, it in enumerate(spot["items"]):
    rr = hr + 1 + i
    ws.cell(row=rr, column=1, value=it.get("id", i)).alignment = TOP
    ws.cell(row=rr, column=2, value=it["definition"]).alignment = WRAP
    for c in (3, 4, 5):
        cell = ws.cell(row=rr, column=c); cell.fill = RESP; cell.border = THIN; cell.alignment = WRAP
    ws.row_dimensions[rr].height = 34
ws.freeze_panes = f"A{hr+1}"
widths(ws, {"A": 5, "B": 58, "C": 30, "D": 28, "E": 34})

# ---------------- Tab 3: knull KEY ----------------
ws = wb.create_sheet("knull KEY (open after)")
ws.cell(row=1, column=1, value="ANSWER KEY — open only AFTER recording all 10 style verdicts on the previous tab.").font = Font(bold=True, size=12, color="7F6000")
ak = spot.get("answer_key")
ws.cell(row=3, column=1, value="id").font = SUB
ws.cell(row=3, column=2, value="key").font = SUB
if isinstance(ak, dict):
    for r0, (k, v) in enumerate(ak.items(), 4):
        ws.cell(row=r0, column=1, value=str(k)).alignment = TOP
        ws.cell(row=r0, column=2, value=json.dumps(v) if not isinstance(v, str) else v).alignment = WRAP
elif isinstance(ak, list):
    for r0, v in enumerate(ak, 4):
        ws.cell(row=r0, column=1, value=r0 - 4).alignment = TOP
        ws.cell(row=r0, column=2, value=json.dumps(v) if not isinstance(v, str) else v).alignment = WRAP
widths(ws, {"A": 8, "B": 80})

# ---------------- Tab 4: OPTIONAL g3 reconcile ----------------
spec = json.load(open(f"{REPO}/human-eval/human-eval-spec.json"))
g3 = next(t for t in spec["tasks"] if t["experiment"] == "g3")
sample = g3["items"][:15]
ws = wb.create_sheet("2. OPTIONAL g3 reconcile")
ws.sheet_view.showGridLines = False
ws.cell(row=1, column=1, value="OPTIONAL — g3 necessity/sufficiency reconciliation sample (15 of 200). Blocks nothing; validates the proxy gold.").font = Font(bold=True, size=12, color="7F6000")
ws.cell(row=2, column=1, value="For each situation: Pass A — is the target claim true under your ordinary use of the key word? Pass B — does the situation satisfy EVERY listed condition? ('yes' only if all hold.) The g3 verdict currently rests on a GPT-5.6/Fable proxy; your labels let me report agreement and re-run if they differ.").alignment = WRAP
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
ws.row_dimensions[2].height = 60
hdr = ["#", "Situation", "Pass A — target claim", "Pass A: true / false / unclear",
       "Pass B — conditions (all must hold)", "Pass B: yes / no / unclear"]
hr = 4
for c, t in enumerate(hdr, 1):
    ws.cell(row=hr, column=c, value=t)
style_header(ws, hr, len(hdr))
for i, it in enumerate(sample):
    rr = hr + 1 + i
    conds = it.get("pass_b", {}).get("conditions", [])
    cond_txt = "\n".join(f"{c.get('cid')}: {c.get('text')}" for c in conds)
    ws.cell(row=rr, column=1, value=it.get("instance_id", i)).alignment = TOP
    ws.cell(row=rr, column=2, value=it.get("text", "")).alignment = WRAP
    ws.cell(row=rr, column=3, value=it.get("pass_a", {}).get("target", "")).alignment = WRAP
    ws.cell(row=rr, column=4).fill = RESP; ws.cell(row=rr, column=4).border = THIN
    ws.cell(row=rr, column=5, value=cond_txt).alignment = WRAP
    ws.cell(row=rr, column=6).fill = RESP; ws.cell(row=rr, column=6).border = THIN
    ws.row_dimensions[rr].height = 60
ws.freeze_panes = f"A{hr+1}"
widths(ws, {"A": 12, "B": 50, "C": 30, "D": 18, "E": 40, "F": 18})

out = f"{REPO}/human-eval/kernel-of-truth-human-review-2026-07-13.xlsx"
wb.save(out)
print("wrote", out)
print("tabs:", wb.sheetnames)
