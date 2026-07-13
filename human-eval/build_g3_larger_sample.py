#!/usr/bin/env python3
"""Build a concept-stratified g3 human Pass A + Pass B sample (~4 per concept) to resolve
whether the registered g3 FAIL (necessity violations, LLM-judged) holds on human gold.

The 15-item #32 sample was end/begin-only (6.7% necessity, below the 0.10 bar); both LLM judges
(Haiku, Opus) ran ~20% and stricter than the human on it. This larger, representative sample lets
us estimate the human concordant necessity rate across ALL g3 concepts.
"""
import json, re, collections
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

REPO = "/home/ec2-user/css/kernel/kernel-of-truth"
PER_CONCEPT = 4

d = json.load(open(f"{REPO}/human-eval/human-eval-spec.json"))
g3 = next(t for t in d["tasks"] if t["experiment"] == "g3")
by_concept = collections.OrderedDict()
for it in g3["items"]:
    m = re.match(r"g3-([a-z]+)-\d+", it.get("instance_id", ""))
    c = m.group(1) if m else "other"
    by_concept.setdefault(c, []).append(it)

sample = []
for c, items in by_concept.items():
    sample.extend(items[:PER_CONCEPT])

H = Font(bold=True, size=12, color="FFFFFF"); HFILL = PatternFill("solid", fgColor="2F5597")
WRAP = Alignment(wrap_text=True, vertical="top"); TOP = Alignment(vertical="top")
THIN = Border(*[Side(style="thin", color="D0D0D0")] * 4); RESP = PatternFill("solid", fgColor="FFF2CC")
SUB = Font(bold=True, size=11)

wb = Workbook()
# Tab 1: overview
ws = wb.active; ws.title = "READ ME FIRST"; ws.sheet_view.showGridLines = False
rows = [
    ("g3 necessity/sufficiency — larger human sample", 15, True, "1F3864"),
    (f"{len(sample)} situations across {len(by_concept)} concepts ({PER_CONCEPT} per concept). ~1–1.5 hours.", 11, False, None),
    ("", 11, False, None),
    ("Why: the g3 experiment tests whether the kernel's decomposition of a concept (e.g. what it means for", 11, False, None),
    ("something to 'end' or 'begin') states conditions that are NECESSARY for the ordinary claim. A registered", 11, False, None),
    ("FAIL rests on LLM judges flagging ~20% 'necessity violations' (the ordinary claim is TRUE but not every", 11, False, None),
    ("listed condition holds). A 15-item pilot you did suggested the LLMs may over-flag vs a human — but it only", 11, False, None),
    ("covered two concepts. This representative sample settles it: if your necessity-violation rate is below ~10%,", 11, False, None),
    ("g3 may flip FAIL→not-FAIL on human gold (a positive for the kernel); if ~20%, the FAIL is confirmed.", 11, False, None),
    ("", 11, False, None),
    ("For each situation, two independent judgements:", 12, True, "C00000"),
    ("  • Pass A — is the target claim TRUE under your ordinary use of the key word? [true / false / unclear]", 11, False, None),
    ("  • Pass B — does the situation satisfy EVERY listed condition? 'yes' ONLY if all hold; else 'no'. [yes / no / unclear]", 11, False, None),
    ("Judge Pass A and Pass B independently — don't let one drive the other. A 'necessity violation' is exactly", 11, False, None),
    ("Pass A = true AND Pass B = no, so those two columns are what matter.", 11, False, None),
]
r = 1
for text, size, bold, color in rows:
    c = ws.cell(row=r, column=1, value=text); c.font = Font(bold=bold, size=size, color=(color or "000000")); r += 1
ws.column_dimensions["A"].width = 110

# Tab 2: the items
ws = wb.create_sheet("g3 sample"); ws.sheet_view.showGridLines = False
hdr = ["#", "concept", "Situation", "Pass A — target claim", "Pass A: true / false / unclear",
       "Pass B — conditions (ALL must hold)", "Pass B: yes / no / unclear", "Notes"]
for ci, t in enumerate(hdr, 1):
    cell = ws.cell(row=1, column=ci, value=t); cell.font = H; cell.fill = HFILL; cell.alignment = WRAP
for i, it in enumerate(sample):
    rr = i + 2
    m = re.match(r"g3-([a-z]+)-\d+", it.get("instance_id", "")); conc = m.group(1) if m else ""
    conds = it.get("pass_b", {}).get("conditions", [])
    cond_txt = "\n".join(f"{c.get('cid')}: {c.get('text')}" for c in conds)
    ws.cell(row=rr, column=1, value=it.get("instance_id", i)).alignment = TOP
    ws.cell(row=rr, column=2, value=conc).alignment = TOP
    ws.cell(row=rr, column=3, value=it.get("text", "")).alignment = WRAP
    ws.cell(row=rr, column=4, value=it.get("pass_a", {}).get("target", "")).alignment = WRAP
    ws.cell(row=rr, column=5).fill = RESP; ws.cell(row=rr, column=5).border = THIN
    ws.cell(row=rr, column=6, value=cond_txt).alignment = WRAP
    ws.cell(row=rr, column=7).fill = RESP; ws.cell(row=rr, column=7).border = THIN
    ws.cell(row=rr, column=8).fill = RESP; ws.cell(row=rr, column=8).border = THIN
    ws.row_dimensions[rr].height = 66
ws.freeze_panes = "A2"
for col, w in {"A": 13, "B": 11, "C": 52, "D": 30, "E": 16, "F": 42, "G": 16, "H": 22}.items():
    ws.column_dimensions[col].width = w

out = f"{REPO}/human-eval/g3-larger-human-sample-2026-07-13.xlsx"
wb.save(out)
print("wrote", out, "| items:", len(sample), "| concepts:", len(by_concept))
