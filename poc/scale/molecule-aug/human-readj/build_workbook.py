#!/usr/bin/env python3
"""Build the s5-human-readj adjudicator WORKBOOK (.xlsx) from the frozen packet.

Reads the materialized, pinned packet under data/s5-human-readj/ (items/*.txt in the
pinned order.json order, rubric.md) and renders a self-contained fillable workbook:
an Instructions/Rubric tab + an Adjudication tab (95 items in the fixed shuffled order,
each with the item text + a verdict dropdown {FAITHFUL,LOSSY,CANNOT-SAY} + the mandatory
criterial-feature audit + notes). Each of the two independent adjudicators fills their
OWN copy. Arm-blind by construction (the packet never carries arm/provenance).
"""
import json, pathlib
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

BASE = pathlib.Path(__file__).resolve().parents[4] / "data/s5-human-readj"
# Output OUTSIDE the pinned corpus dir (writing into BASE would drift its corpus hash).
OUT = pathlib.Path(__file__).resolve().parent / "s5-human-readj-adjudication.xlsx"

order = json.loads((BASE / "order.json").read_text())
order = order["order"] if isinstance(order, dict) and "order" in order else order
rubric = (BASE / "rubric.md").read_text()

wb = Workbook()
HEAD = PatternFill("solid", fgColor="1F3864"); HEADF = Font(color="FFFFFF", bold=True)
YELL = PatternFill("solid", fgColor="FFF2CC"); wrap = Alignment(wrap_text=True, vertical="top")
thin = Border(*[Side(style="thin", color="BFBFBF")] * 4)

# ---- Instructions / Rubric ----
ins = wb.active; ins.title = "Instructions"; ins.column_dimensions["A"].width = 118
for i, line in enumerate(rubric.splitlines(), start=1):
    c = ins.cell(row=i, column=1, value=line.replace("> ", "").replace(">", ""))
    c.alignment = wrap
    if line.startswith("#"): c.font = Font(bold=True, size=13 if line.startswith("# ") else 11)

# ---- Adjudication ----
adj = wb.create_sheet("Adjudication")
cols = ["#", "Item ID", "Item (concept + candidate explication)", "VERDICT",
        "Criterial-feature audit (required)", "Notes (optional)"]
widths = [5, 12, 88, 18, 46, 30]
for i, (h, w) in enumerate(zip(cols, widths), start=1):
    c = adj.cell(row=1, column=i, value=h); c.fill = HEAD; c.font = HEADF
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    adj.column_dimensions[get_column_letter(i)].width = w
adj.freeze_panes = "A2"

for idx, item_id in enumerate(order, start=1):
    txt_path = BASE / "items" / (item_id + ".txt")
    text = txt_path.read_text() if txt_path.exists() else "(missing item file)"
    r = idx + 1
    for ci, v in enumerate([idx, item_id, text, "", "", ""], start=1):
        c = adj.cell(row=r, column=ci, value=v); c.alignment = wrap; c.border = thin
        if ci in (4, 5): c.fill = YELL
    adj.row_dimensions[r].height = 120

dv = DataValidation(type="list", formula1='"FAITHFUL,LOSSY,CANNOT-SAY"', allow_blank=True, showDropDown=False)
dv.prompt = "Pick your verdict"; adj.add_data_validation(dv)
for r in range(2, len(order) + 2):
    dv.add(adj.cell(row=r, column=4))

wb.save(OUT)
print(f"built {OUT.name}: {len(order)} items, arm-blind, in the pinned shuffled order")
