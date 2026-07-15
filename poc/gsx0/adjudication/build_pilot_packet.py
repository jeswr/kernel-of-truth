#!/usr/bin/env python3
"""Build the gsx0 STAGE-0 pilot BLIND adjudication packet for judge-1 (kernel-naive human).

Produces, under poc/gsx0/adjudication/pilot-first60/:
  - gsx0-pilot-adjudication.xlsx   (self-contained workbook: Instructions + Adjudication)
  - preview.csv                    (GitHub-renderable table of what the human sees)
  - token-map.json                 (opaque token -> item_id; coordinator-only, NOT shown to judge)

BLINDING (PROTOCOL.md §4.2): the human sees ONLY the question text and (for MCQ) the option
texts. Never: urn, record_path, membership gold `answer`, `type` label, provenance, which option
is the kernel's own gloss. Presentation order is a deterministic seed-derived shuffle
(`dadjt/1|judge-1|20260710`); labels are keyed by item id so order never affects the stored gold.
Item selection = the FIRST 60 items of d-qa-t-plain in pinned rank order (design §5, STAGE 0).
Escape hatch (§4.3): claims -> yes/no/cannot-say ; MCQ -> A/B/C/D or NONE-or-cannot-say.
"""
import json, hashlib, csv, pathlib
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path(__file__).resolve().parents[3]
ITEMS = ROOT / "data/d-qa-t-plain/items/covered.jsonl"
OUT = ROOT / "poc/gsx0/adjudication/pilot-first60"
OUT.mkdir(parents=True, exist_ok=True)
SEED = "dadjt/1|judge-1|20260710"   # PROTOCOL.md §4.2 pinned per-judge permutation seed

# --- load first 60 by pinned rank -------------------------------------------------
rows = [json.loads(l) for l in ITEMS.open()]
rows.sort(key=lambda r: r["rank"])
first60 = rows[:60]
assert len(first60) == 60 and [r["rank"] for r in first60] == list(range(60)), "rank 0..59 expected"

# --- deterministic blind shuffle (seed-derived, reproducible, carries 0 bits about answers) ---
def sortkey(r):
    return hashlib.sha256(f"{SEED}|{r['id']}".encode()).hexdigest()
shuffled = sorted(first60, key=sortkey)
tokens = [f"P{ i+1:02d}" for i in range(60)]

# --- workbook ---------------------------------------------------------------------
wb = Workbook()
HEAD = PatternFill("solid", fgColor="1F3864"); HEADF = Font(color="FFFFFF", bold=True, size=11)
SUB  = PatternFill("solid", fgColor="D6DCE5")
YELL = PatternFill("solid", fgColor="FFF2CC")
wrap = Alignment(wrap_text=True, vertical="top")
thin = Border(*[Side(style="thin", color="BFBFBF")]*4)

# ---- Sheet 1: Instructions ----
ins = wb.active; ins.title = "Instructions"
ins.column_dimensions["A"].width = 112
def line(txt, bold=False, size=11, fill=None):
    r = ins.max_row + 1 if ins["A1"].value is not None else 1
    c = ins.cell(row=r, column=1, value=txt)
    c.alignment = wrap; c.font = Font(bold=bold, size=size)
    if fill: c.fill = fill
    return r
line("gsx0 STAGE-0 pilot — blind external adjudication (judge-1)", bold=True, size=14, fill=SUB)
for t in [
 "",
 "WHO SHOULD DO THIS.  judge-1 MUST be a person who has NEVER read any kernel-v0 or molecules-v0",
 "record — i.e. kernel-naive.  Your answers are the SOLE gold standard for this pilot, so if you have",
 "seen the project's concept definitions, please pass this to someone who has not.  Use only ordinary,",
 "everyday understanding of the words — do NOT look anything up.",
 "",
 "THE TASK.  60 short items on the 'Adjudication' tab.  Two kinds:",
 "  • CLAIM items — 'According to the definition of X, is <S> true of X?'  Answer yes / no / cannot-say.",
 "  • CHOICE items — a definition or a word plus four options A–D.  Pick the best option, or",
 "    'NONE-or-cannot-say'.",
 "Put your answer in the yellow YOUR ANSWER column (a dropdown).  ~15–20 minutes total.  Judge each item",
 "ON ITS OWN — the same sentence can appear under different words; never reuse an earlier answer.",
 "",
 "HOW TO DECIDE (the standards this pilot is scored against):",
 "  S1  CHOICE = fit + identification.  An option is correct iff (a) everything it says fits the word as",
 "      ordinarily understood — read each clause as a typical case, honouring hedges ('can', 'often',",
 "      'some'); AND (b) taken as a whole it says what the word MEANS (picks out this concept, not a",
 "      near one).  Extra TRUE facts beyond the core do not disqualify (surplus truth is fine).  Any",
 "      clause that is FALSE of the word disqualifies.  A pile of true facts that never says what the",
 "      thing is → not correct.",
 "  S2  If more than one option passes S1, pick the one that best/most specifically gives the meaning of",
 "      the ASKED word itself.  Choose NONE-or-cannot-say only if no option passes, or you cannot decide.",
 "  S3  WORD-match (given a definition, pick the word) is S1 in reverse.  A parenthetical like",
 "      'find (X finds Y)' only tells you which sense is meant — ignore any unfamiliar notation inside it;",
 "      it is never itself a reason to answer NONE.",
 "  S4  CLAIMS at the generic standard.  yes = S holds in the normal case (exceptions don't make it",
 "      false), or S's own hedge holds.  no = S misdescribes the word, OR S has nothing to do with what",
 "      the word means (including statements so generic they say nothing in particular about it).",
 "      cannot-say = ONLY genuine inability to understand/decide — never a soft 'no', never just because",
 "      the wording is odd or only sometimes true.",
 "  S5  Fragments & participants.  Claims are fragments of a longer description: unresolved 'this someone /",
 "      it / these parts', stray quote marks, and '[bracketed]' notes are read charitably as pieces of a",
 "      description of the word's normal scenario — judge whether the fragment could belong to describing",
 "      the word.  Letters X/Y name the participants named with the word ('break (X breaks Y)': X breaks,",
 "      Y gets broken); if the X/Y roles match nothing in the word's meaning → no.",
 "  S6  Register immunity.  The deliberately plain wording is never a reason for NONE/no/cannot-say:",
 "      'something of kind K' = 'a K' ('a something of kind event' = 'an event').  Odd grammar is",
 "      simplification, not a trick.",
 "  S7  Independence.  Judge each item only against its own word — never by pattern or by recalling an",
 "      earlier item.",
 "",
 "WHEN DONE.  Save the file and return it (re-upload where you got it, or send it back).  Do not add,",
 "reorder, or delete rows.  Thank you!",
]:
    line(t)

# ---- Sheet 2: Adjudication ----
adj = wb.create_sheet("Adjudication")
cols = ["Item", "Question", "Option A", "Option B", "Option C", "Option D",
        "YOUR ANSWER", "Notes (optional)"]
widths = [8, 66, 30, 30, 30, 30, 20, 34]
for i, (h, w) in enumerate(zip(cols, widths), start=1):
    c = adj.cell(row=1, column=i, value=h); c.fill = HEAD; c.font = HEADF
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    adj.column_dimensions[get_column_letter(i)].width = w
adj.freeze_panes = "A2"

claim_rows, mcq_rows = [], []
tokenmap = {}
preview = []
for idx, (tok, it) in enumerate(zip(tokens, shuffled), start=2):
    tokenmap[tok] = it["id"]
    is_mcq = bool(it["options"])
    optvals = ["", "", "", ""]
    if is_mcq:
        for o in it["options"]:
            k = o["key"]
            if k in ("A", "B", "C", "D"):
                optvals["ABCD".index(k)] = o["text"]
        mcq_rows.append(idx)
    else:
        claim_rows.append(idx)
    vals = [tok, it["question"], *optvals, "", ""]
    for ci, v in enumerate(vals, start=1):
        c = adj.cell(row=idx, column=ci, value=v)
        c.alignment = wrap; c.border = thin
        if ci == 7: c.fill = YELL
    adj.row_dimensions[idx].height = 58 if is_mcq else 34
    preview.append({"Item": tok, "Kind": "CHOICE" if is_mcq else "CLAIM",
                    "Question": it["question"],
                    "A": optvals[0], "B": optvals[1], "C": optvals[2], "D": optvals[3]})

def add_dv(formula, rowlist):
    if not rowlist: return
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
    dv.error = "Pick a value from the dropdown."; dv.prompt = "Choose your answer"
    adj.add_data_validation(dv)
    for r in rowlist:
        dv.add(adj.cell(row=r, column=7))
add_dv('"yes,no,cannot-say"', claim_rows)
add_dv('"A,B,C,D,NONE-or-cannot-say"', mcq_rows)

xlsx = OUT / "gsx0-pilot-adjudication.xlsx"
wb.save(xlsx)
(OUT / "token-map.json").write_text(json.dumps(
    {"seed": SEED, "note": "coordinator-only; NOT shown to judge-1", "map": tokenmap}, indent=1))
with (OUT / "preview.csv").open("w", newline="") as f:
    w = csv.DictWriter(f, fieldnames=["Item", "Kind", "Question", "A", "B", "C", "D"])
    w.writeheader(); w.writerows(preview)

print(f"built {xlsx.name}: 60 items ({len(claim_rows)} claim, {len(mcq_rows)} choice)")
print(f"  seed={SEED}")
print(f"  token-map.json + preview.csv written to {OUT.relative_to(ROOT)}")
